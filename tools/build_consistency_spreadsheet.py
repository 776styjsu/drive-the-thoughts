#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
# Copyright (c) 2026 NVIDIA Corporation

"""Build a per-clip spreadsheet joining CoT consistency judgments with a benchmark.

Each row is one CoT-trajectory pair (clip). Columns combine ground-truth
annotations (consistency, CoT reliability, scene labels), the judge's
prediction (label, score, justification), correctness vs. the benchmark, and
trajectory summary stats, so patterns in agreements/disagreements can be
filtered and sorted in any spreadsheet tool.

Accepts either an accuracy file produced by tools/check_consistency_accuracy.py
(*.accuracy.json) or a raw judge output (*.cot_consistency*.json). When given
an accuracy file, the raw judge file it references is also loaded (if present)
to pull judge justifications and trajectory stats.

Examples:
    python3 tools/build_consistency_spreadsheet.py \
        --consistency-file benchmark_expanded_100.cot_consistency_gpt55.accuracy.json \
        --benchmark-file benchmark_expanded_100.json

    # Write an .xlsx instead of .csv (requires openpyxl, e.g. via `uv run`):
    uv run python tools/build_consistency_spreadsheet.py \
        --consistency-file benchmark_expanded_100.cot_consistency_gpt55.accuracy.json \
        --benchmark-file benchmark_expanded_100.json \
        --output benchmark_expanded_100.pairs.xlsx
"""

import argparse
import ast
import csv
import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

# Match tools/check_consistency_accuracy.py: score > 2 => consistent.
LLM_INCONSISTENT_THRESHOLD = 2

SIMPLE_COLUMNS = [
    "clip_id",
    "chain_of_thought",
    "gt_consistent",
    "predicted_consistent",
    "score",
    "judge_correct",
    "cot_reliable",
    "cot_reliability_justification",
    "gt_justification",
    "judge_justification"
]

COLUMNS = [
    "clip_id",
    "chain_of_thought",
    "gt_consistent",
    "predicted_consistent",
    "score",
    "judge_correct",
    "disagreement_type",
    "cot_reliable",
    "cot_unreliability_taxonomy",
    "cot_reliability_justification",
    "gt_inconsistency_subtypes",
    "gt_justification",
    "judge_justification",
    "planned_trajectory_safe",
    "planned_trajectory_law_compliant",
    "behavior",
    "layout",
    "lighting",
    "road_types",
    "traffic_density",
    "weather",
    "vrus",
    "gt_longitudinal_decision",
    "gt_lateral_decision",
    "meta_action_longitudinal",
    "meta_action_lateral",
    "duration_s",
    "total_path_length_m",
    "final_lateral_m",
    "max_lateral_m",
    "mean_speed_ms",
    "max_speed_ms",
    "min_speed_ms",
    "mean_accel_ms2",
    "note",
    "rollout_mp4",
]


def load_json_file(file_path: Union[str, Path]) -> Union[dict, list]:
    with Path(file_path).open("r", encoding="utf-8") as f:
        return json.load(f)


def clean_cot(raw: Any) -> str:
    """Render chain_of_thought (often a stringified Python list) as readable text."""
    if raw is None:
        return ""
    if isinstance(raw, list):
        return " | ".join(str(x) for x in raw)
    text = str(raw).strip()
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = ast.literal_eval(text)
            if isinstance(parsed, (list, tuple)):
                return " | ".join(str(x) for x in parsed)
        except (ValueError, SyntaxError):
            pass
    return text


def join_list(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(x) for x in value)
    return str(value)


def flatten_unreliability_taxonomy(value: Any) -> str:
    """Render the taxonomy dict as its category list, e.g. 'scene_entity_hallucination'."""
    if not isinstance(value, dict):
        return join_list(value)
    categories = value.get("categories") or []
    primary = value.get("primary_category")
    if primary and primary not in categories:
        categories = [primary] + list(categories)
    return "; ".join(str(c) for c in categories)


def cot_reliability_flag(entry: dict) -> Any:
    """CoT reliability from a benchmark entry, supporting both on-disk schemas.

    - flat ``cot_reliable`` (bool/str), used by benchmark_expanded_*.json
    - nested ``cot_reliability.reliable`` (bool), used by benchmark.json

    Returns True/False when a signal is present, else None.
    """
    nested = entry.get("cot_reliability")
    value = (
        nested.get("reliable")
        if isinstance(nested, dict) and "reliable" in nested
        else entry.get("cot_reliable")
    )
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        text = value.strip().lower()
        if text in {"true", "reliable", "yes", "1"}:
            return True
        if text in {"false", "unreliable", "no", "0"}:
            return False
    return None


def reliability_justification(entry: dict) -> Any:
    """Reliability justification from either schema (flat or nested)."""
    nested = entry.get("cot_reliability")
    if isinstance(nested, dict) and nested.get("justification") is not None:
        return nested.get("justification")
    return entry.get("cot_reliability_justification")


def unreliability_taxonomy(entry: dict) -> Any:
    """Taxonomy value (renderable by ``flatten_unreliability_taxonomy``) from either schema."""
    nested = entry.get("cot_reliability")
    if isinstance(nested, dict) and (
        nested.get("unreliability_categories")
        or nested.get("primary_unreliability_category")
    ):
        return {
            "categories": nested.get("unreliability_categories") or [],
            "primary_category": nested.get("primary_unreliability_category"),
        }
    return entry.get("cot_unreliability_taxonomy")


def flatten_inconsistency_subtypes(value: Any) -> str:
    """Render the subtype dict as the subtypes marked present, e.g. 'magnitude_mismatch'."""
    if not isinstance(value, dict):
        return join_list(value)
    present = [
        name
        for name, info in value.items()
        if isinstance(info, dict) and str(info.get("present", "")).lower() == "yes"
    ]
    return "; ".join(sorted(present))


def index_predictions(consistency_data: dict) -> Dict[str, dict]:
    """Map clip_id -> prediction record from an accuracy file's matches/mismatches."""
    predictions: Dict[str, dict] = {}
    for group, correct in (("matches", True), ("mismatches", False)):
        for entry in consistency_data.get(group, []):
            clip_id = entry.get("clip_id")
            if clip_id:
                predictions[clip_id] = {**entry, "judge_correct": correct}
    return predictions


def index_judge_results(judge_data: dict) -> Dict[str, dict]:
    """Map clip_id -> raw judge result (justification, trajectory stats, score)."""
    results: Dict[str, dict] = {}
    for entry in judge_data.get("results", []):
        clip_id = entry.get("clip_id")
        if clip_id:
            results[clip_id] = entry
    return results


def predictions_from_judge(judge_results: Dict[str, dict]) -> Dict[str, dict]:
    """Derive prediction records from a raw judge file when no accuracy file is given."""
    predictions: Dict[str, dict] = {}
    for clip_id, entry in judge_results.items():
        evaluation = entry.get("evaluation", {}).get("cot_output_alignment", {})
        score = evaluation.get("score")
        if score is None:
            continue
        predictions[clip_id] = {
            "clip_id": clip_id,
            "predicted_is_consistent": score > LLM_INCONSISTENT_THRESHOLD,
            "prediction_score": score,
        }
    return predictions


def resolve_referenced_judge_file(
    accuracy_path: Path, referenced: Optional[str]
) -> Optional[Path]:
    if not referenced:
        return None
    candidates = [Path(referenced), accuracy_path.parent / referenced]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def build_rows(
    benchmark_entries: List[dict],
    predictions: Dict[str, dict],
    judge_results: Dict[str, dict],
) -> List[Dict[str, Any]]:
    rows = []
    for entry in benchmark_entries:
        clip_id = entry.get("clip_id", "")
        prediction = predictions.get(clip_id, {})
        judge = judge_results.get(clip_id, {})

        labels = entry.get("labels", {})
        scene = labels.get("nurec_scene", {})
        cot_label = labels.get("cot_decision_label", {})
        decision = cot_label.get("high_level_decision", {})
        meta_action = cot_label.get("atomic_meta_action_hint", {})

        evaluation = judge.get("evaluation", {}).get("cot_output_alignment", {})
        stats = judge.get("trajectory_features", {}).get("summary_stats", {})

        gt_consistent = entry.get("cot_action_consistency")
        predicted_consistent = prediction.get("predicted_is_consistent")
        judge_correct = prediction.get("judge_correct")
        if judge_correct is None and predicted_consistent is not None:
            judge_correct = predicted_consistent == gt_consistent

        rows.append(
            {
                "clip_id": clip_id,
                "chain_of_thought": clean_cot(entry.get("chain_of_thought")),
                "gt_consistent": gt_consistent,
                "predicted_consistent": predicted_consistent,
                "score": prediction.get("prediction_score", evaluation.get("score")),
                "judge_correct": judge_correct,
                "disagreement_type": prediction.get("disagreement_type", ""),
                "cot_reliable": cot_reliability_flag(entry),
                "cot_unreliability_taxonomy": flatten_unreliability_taxonomy(
                    unreliability_taxonomy(entry)
                ),
                "cot_reliability_justification": join_list(
                    reliability_justification(entry)
                ),
                "gt_inconsistency_subtypes": flatten_inconsistency_subtypes(
                    entry.get("inconsistency_subtypes")
                ),
                "gt_justification": join_list(entry.get("justification")),
                "judge_justification": evaluation.get("justification", ""),
                "planned_trajectory_safe": entry.get("planned_trajectory_safe"),
                "planned_trajectory_law_compliant": entry.get(
                    "planned_trajectory_law_compliant"
                ),
                "behavior": join_list(scene.get("behavior")),
                "layout": join_list(scene.get("layout")),
                "lighting": join_list(scene.get("lighting")),
                "road_types": join_list(scene.get("road_types")),
                "traffic_density": join_list(scene.get("traffic_density")),
                "weather": join_list(scene.get("weather")),
                "vrus": scene.get("vrus"),
                "gt_longitudinal_decision": decision.get("longitudinal", ""),
                "gt_lateral_decision": decision.get("lateral", ""),
                "meta_action_longitudinal": join_list(
                    meta_action.get("longitudinal_sequence")
                    or meta_action.get("longitudinal")
                ),
                "meta_action_lateral": join_list(
                    meta_action.get("lateral_sequence") or meta_action.get("lateral")
                ),
                "duration_s": stats.get("duration_s"),
                "total_path_length_m": stats.get("total_path_length_m"),
                "final_lateral_m": stats.get("final_lateral_m"),
                "max_lateral_m": stats.get("max_lateral_m"),
                "mean_speed_ms": stats.get("mean_speed_ms"),
                "max_speed_ms": stats.get("max_speed_ms"),
                "min_speed_ms": stats.get("min_speed_ms"),
                "mean_accel_ms2": stats.get("mean_accel_ms2"),
                "note": join_list(entry.get("note")),
                "rollout_mp4": entry.get("rollout_mp4", ""),
            }
        )
    return rows


def write_csv(
    rows: List[Dict[str, Any]], output_path: Path, columns: List[str]
) -> None:
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(
    rows: List[Dict[str, Any]], output_path: Path, columns: List[str]
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "cot_trajectory_pairs"

    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    mismatch_fill = PatternFill(
        start_color="FFF4CCCC", end_color="FFF4CCCC", fill_type="solid"
    )
    no_prediction_fill = PatternFill(
        start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid"
    )
    for row in rows:
        ws.append([row.get(col) for col in columns])
        fill = None
        if row.get("judge_correct") is False:
            fill = mismatch_fill
        elif row.get("predicted_consistent") is None:
            fill = no_prediction_fill
        if fill is not None:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    wide_columns = {
        "chain_of_thought",
        "cot_reliability_justification",
        "gt_justification",
        "judge_justification",
        "rollout_mp4",
    }
    for idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 60 if col in wide_columns else 18

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "--consistency-file",
        required=True,
        type=Path,
        help="Accuracy file (*.accuracy.json) or raw judge output (*.cot_consistency*.json)",
    )
    parser.add_argument(
        "--benchmark-file",
        required=True,
        type=Path,
        help="Benchmark JSON with ground-truth annotations",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output path (.csv or .xlsx). Default: <consistency-file stem>.pairs.csv",
    )
    parser.add_argument(
        "--full",
        action="store_true",
        help="Include all columns (default: simple view with clip_id, "
        "chain_of_thought, gt/predicted consistency, score, judge_correct, "
        "cot_reliable)",
    )
    args = parser.parse_args()
    columns = COLUMNS if args.full else SIMPLE_COLUMNS

    consistency_data = load_json_file(args.consistency_file)
    benchmark_data = load_json_file(args.benchmark_file)
    benchmark_entries = (
        benchmark_data if isinstance(benchmark_data, list) else benchmark_data.get("results", [])
    )

    judge_results: Dict[str, dict] = {}
    if isinstance(consistency_data, dict) and "matches" in consistency_data:
        predictions = index_predictions(consistency_data)
        judge_path = resolve_referenced_judge_file(
            args.consistency_file, consistency_data.get("consistency_file")
        )
        if judge_path is not None:
            judge_results = index_judge_results(load_json_file(judge_path))
            print(f"Loaded judge justifications from {judge_path}")
        else:
            print(
                "Warning: raw judge file not found; judge_justification and "
                "trajectory stats columns will be empty"
            )
    elif isinstance(consistency_data, dict) and "results" in consistency_data:
        judge_results = index_judge_results(consistency_data)
        predictions = predictions_from_judge(judge_results)
    else:
        raise SystemExit(
            "Unrecognized consistency file format: expected an accuracy file "
            "(with 'matches'/'mismatches') or a raw judge output (with 'results')"
        )

    rows = build_rows(benchmark_entries, predictions, judge_results)

    default_name = args.consistency_file.name
    if default_name.endswith(".json"):
        default_name = default_name[: -len(".json")]
    output_path = args.output or args.consistency_file.parent / (
        default_name + ".pairs.csv"
    )
    if output_path.suffix.lower() == ".xlsx":
        write_xlsx(rows, output_path, columns)
    else:
        write_csv(rows, output_path, columns)

    n_predicted = sum(1 for r in rows if r["predicted_consistent"] is not None)
    n_wrong = sum(1 for r in rows if r["judge_correct"] is False)
    print(
        f"Wrote {len(rows)} CoT-trajectory pairs to {output_path} "
        f"({n_predicted} with predictions, {n_wrong} judge mismatches)"
    )


if __name__ == "__main__":
    main()
